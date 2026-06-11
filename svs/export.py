"""
Excel export helpers.

Two distinct exports:
  build_parser_excel()   — styled workbook from the parser's DataFrame
                           (flags low-confidence cells in amber)
  build_schedule_excel() — multi-sheet workbook with summary + per-day timelines
"""

from __future__ import annotations
import io

import pandas as pd

from svs.config import DAY_CONFIG, DAY4_VP_CONFIG
from svs.scheduler import build_timeline_df, build_unassigned_df, build_summary_df


# ── Parser export ─────────────────────────────────────────────────────────────

def build_parser_excel(df_export: pd.DataFrame, flagged_cells: set[tuple[str, str]]) -> bytes:
    """
    Build a styled openpyxl workbook from the parser DataFrame and return bytes.

    flagged_cells: set of (user_id_str, field_name) tuples that should be
                   highlighted amber (low/medium confidence).

    Raises ImportError with a clear message if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export.  "
            "Run: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SvS Data"

    HEADER_FILL  = PatternFill("solid", fgColor="1A2535")
    FLAGGED_FILL = PatternFill("solid", fgColor="1E1608")
    EMPTY_FILL   = PatternFill("solid", fgColor="130E0E")
    THIN         = Side(style="thin", color="1E2533")
    BORDER       = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    cols    = list(df_export.columns)
    uid_col = cols.index("User ID") if "User ID" in cols else None

    # Header row
    for ci, col in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color="7AABDC", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[1].height = 22

    # Data rows
    for ri, (_, row) in enumerate(df_export.iterrows(), 2):
        uid = str(row["User ID"]) if uid_col is not None else None
        for ci, col in enumerate(cols, 1):
            val  = row[col]
            cell = ws.cell(row=ri, column=ci, value=None if val in ("", "—") else val)
            cell.alignment = Alignment(vertical="center")
            cell.border    = BORDER
            if val in ("", "—", None):
                cell.fill = EMPTY_FILL
                cell.font = Font(color="4A5568")
            elif uid and (uid, col) in flagged_cells:
                cell.fill = FLAGGED_FILL
                cell.font = Font(color="D4A855")
            else:
                cell.font = Font(color="C9D1DC")
        ws.row_dimensions[ri].height = 18

    # Auto-fit column widths
    for ci, col in enumerate(cols, 1):
        cell_lengths = [len(str(v)) for v in df_export[col].fillna("")]
        max_len      = max([len(str(col))] + cell_lengths, default=len(str(col)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 42)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Scheduler export ──────────────────────────────────────────────────────────

def build_schedule_excel(
    users: list[dict],
    day_results: list[dict],
    vp4_result: dict | None = None,
) -> io.BytesIO:
    """
    Export scheduler results to a multi-sheet xlsx file.

    Sheets:
      Summary          — one row per player, all days side by side
      Day_N_<label>    — slot timeline for each day
      Unassigned_DayN  — unassigned players for each day (if any)
      Day_4_VP         — Day 4 VP timeline (if vp4_result provided)
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        build_summary_df(users, day_results, vp4_result).to_excel(
            writer, sheet_name="Summary", index=False
        )
        for dr in day_results:
            dc    = next(d for d in DAY_CONFIG if d["day"] == dr["day"])
            sheet = dc["label"].replace(" — ", " ").replace(" ", "_")[:31]
            build_timeline_df(users, dr).to_excel(
                writer, sheet_name=sheet, index=False
            )
            ua = build_unassigned_df(dr)
            if not ua.empty:
                ua.to_excel(
                    writer, sheet_name=f"Unassigned_Day{dc['day']}", index=False
                )

        if vp4_result is not None and (vp4_result["user_slot"] or vp4_result["unassigned"]):
            build_timeline_df(users, vp4_result).to_excel(
                writer, sheet_name="Day_4_VP", index=False
            )
            ua_vp4 = build_unassigned_df(vp4_result)
            if not ua_vp4.empty:
                ua_vp4.to_excel(writer, sheet_name="Unassigned_Day4_VP", index=False)

    buf.seek(0)
    return buf